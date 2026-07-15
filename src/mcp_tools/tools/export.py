"""
Unified export tool.

Replaces 4 individual tools with 1:
- export_data: export_drawing_to_excel, extract_drawing_data,
               export_selected_to_excel, extract_selected_data (4→1)

Two dimensions:
  scope:  "all" or "selected"
  format: "json" or "excel"
"""

import json
import logging
from pathlib import Path

from mcp_tools.decorators import cad_tool, get_current_adapter

logger = logging.getLogger(__name__)


def _set_cell_value_safe(ws, row: int, col: int, value):
    """Write value to cell, handling merged ranges."""
    from openpyxl.cell.cell import MergedCell

    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if (
                merged_range.min_row <= row <= merged_range.max_row
                and merged_range.min_col <= col <= merged_range.max_col
            ):
                ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
                return
    cell.value = value


def _export_excel(adapter, data, filename: str, scope_label: str) -> str:
    """Shared Excel export logic for both all/selected scopes."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from core.config import get_config

    config = get_config()
    output_dir = Path(config.output.directory).expanduser().resolve()

    # Determine export path
    file_obj = Path(filename)
    dir_part = str(file_obj.parent)
    if dir_part and dir_part != ".":
        export_dir = Path(dir_part).expanduser().resolve()
    else:
        export_dir = output_dir

    # Security: verify within output directory
    try:
        export_dir.relative_to(output_dir)
    except ValueError:
        logger.error(
            f"Security: Attempted export outside output directory. "
            f"Requested: {export_dir}, Allowed: {output_dir}"
        )
        return json.dumps({"success": False, "error": "Invalid output directory"}, indent=2)

    export_dir.mkdir(parents=True, exist_ok=True)
    full_filepath = export_dir / file_obj.name

    # Create workbook
    workbook = Workbook()
    worksheet = workbook.active
    if worksheet is None:
        return json.dumps({"success": False, "error": "Failed to create worksheet"}, indent=2)

    worksheet.title = f"{scope_label} Data"

    columns = [
        "Handle",
        "ObjectType",
        "Layer",
        "Color",
        "Length",
        "Area",
        "Radius",
        "Circumference",
        "Name",
    ]

    # Headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(columns, 1):
        try:
            _set_cell_value_safe(worksheet, 1, col_idx, col_name)
            cell = worksheet.cell(row=1, column=col_idx)
            if cell is not None:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        except Exception as e:
            logger.debug(f"Error writing header at column {col_idx}: {e}")

    # Data rows
    numeric_cols = {"Length", "Area", "Radius", "Circumference"}
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            try:
                value = row_data.get(col_name)
                _set_cell_value_safe(worksheet, row_idx, col_idx, value)
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell is not None:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    if col_name in numeric_cols and isinstance(value, (int, float)):
                        cell.number_format = "0.000"
            except Exception as e:
                logger.debug(f"Error writing data at row {row_idx}, col {col_idx}: {e}")

    # Auto-adjust column widths
    for col_idx, col_name in enumerate(columns, 1):
        max_length = len(col_name)
        for row_idx in range(2, len(data) + 2):
            cell_obj = worksheet.cell(row=row_idx, column=col_idx)
            cell_value = str(cell_obj.value or "") if cell_obj is not None else ""
            max_length = max(max_length, len(cell_value))
        col_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

    worksheet.freeze_panes = "A2"
    workbook.save(str(full_filepath))
    logger.info(f"Exported {len(data)} entities to {full_filepath}")

    return json.dumps(
        {
            "success": True,
            "count": len(data),
            "message": f"Exported {len(data)} entities to {filename}",
            "path": str(full_filepath),
        },
        indent=2,
    )


# ========== Tool Registration ==========


def register_export_tools(mcp):
    """Register unified export tool with FastMCP."""

    @cad_tool(mcp, "export_data")
    def export_data(
        scope: str = "all",
        format: str = "json",
        filename: str = "drawing_data.xlsx",
    ) -> str:
        """
        Extract or export drawing data.

        Args:
            scope: What to export.
                   - "all": all entities in the drawing
                   - "selected": only currently selected entities
            format: Output format.
                    - "json": return data as JSON (for analysis by the LLM)
                    - "excel": save to Excel file
            filename: Excel filename (only used when format="excel").
                      Default: "[drawing_name]_data.xlsx" (or "drawing_data.xlsx" if unnamed).
                      Saved to the configured output directory.


        Returns:
            JSON result with entity data or export status

        Examples:
            scope="all", format="json"              → extract all entities as JSON
            scope="all", format="excel"             → export all to Excel
            scope="selected", format="json"         → extract selected as JSON
            scope="selected", format="excel", filename="selection.xlsx"

        Data columns: Handle, ObjectType, Layer, Color, Length, Area,
                      Radius, Circumference, Name
        """
        adapter = get_current_adapter()
        scope_lower = scope.lower()
        format_lower = format.lower()

        # Validate parameters
        if scope_lower not in ("all", "selected"):
            return json.dumps(
                {
                    "success": False,
                    "error": f"Unknown scope '{scope}'. Use: all, selected",
                },
                indent=2,
            )
        if format_lower not in ("json", "excel"):
            return json.dumps(
                {
                    "success": False,
                    "error": f"Unknown format '{format}'. Use: json, excel",
                },
                indent=2,
            )

        if format_lower == "excel" and filename == "drawing_data.xlsx":
            try:
                doc = adapter._get_document("export_data")
                doc_stem = Path(doc.Name).stem if doc and doc.Name else "drawing"
                filename = f"{doc_stem}_data.xlsx"
            except Exception:
                pass

        try:
            # For "selected" scope, check selection first
            if scope_lower == "selected":
                if format_lower == "json":
                    selection_info = adapter.get_selection_info()
                    if selection_info["count"] == 0:
                        return json.dumps(
                            {
                                "success": True,
                                "count": 0,
                                "message": "No entities selected",
                                "entities": [],
                            },
                            indent=2,
                        )
                else:
                    if not adapter.has_selection():
                        return json.dumps(
                            {
                                "success": False,
                                "error": "No entities selected. Select entities first.",
                            },
                            indent=2,
                        )

            # Extract data
            only_selected = scope_lower == "selected"

            if format_lower == "excel" and scope_lower == "all":
                # Use adapter's built-in export for full drawing
                success = adapter.export_to_excel(filename)
                return json.dumps(
                    {
                        "success": success,
                        "message": (f"Exported to {filename}" if success else "Export failed"),
                    },
                    indent=2,
                )

            # Extract data (for json output or selected excel)
            data = adapter.extract_drawing_data(only_selected=only_selected, limit=0)

            if not data:
                return json.dumps(
                    {
                        "success": scope_lower == "all",
                        "count": 0,
                        "message": "No exportable data found",
                        "entities": [] if format_lower == "json" else None,
                    },
                    indent=2,
                )

            # JSON output
            if format_lower == "json":
                result = {
                    "success": True,
                    "count": len(data),
                    "message": f"Extracted data from {len(data)} entities",
                    "entities": data,
                }
                # Add UI metadata for full drawing extraction
                if scope_lower == "all":
                    result["_meta"] = {
                        "ui": {
                            "resourceUri": "ui://multicad/drawing_viewer",
                            "data": {"entities": data},
                        }
                    }
                return json.dumps(result, indent=2)

            # Excel output (selected scope)
            scope_label = "Selected" if only_selected else "Drawing"
            return _export_excel(adapter, data, filename, scope_label)

        except Exception as e:
            logger.error(f"Export/extract failed: {e}")
            return json.dumps(
                {"success": False, "error": str(e), "count": 0},
                indent=2,
            )
