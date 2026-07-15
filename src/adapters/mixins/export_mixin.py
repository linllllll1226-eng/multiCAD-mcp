"""
Export mixin for AutoCAD adapter.

Handles data extraction and Excel export operations.
"""

import logging
import math
from contextlib import contextmanager
from typing import TYPE_CHECKING, Any, Dict, List, Optional

from mcp_tools.constants import COLOR_MAP

logger = logging.getLogger(__name__)


class ExportMixin:
    """Mixin for data extraction and export operations."""

    if TYPE_CHECKING:

        def _validate_connection(self) -> None: ...

        def _get_document(self, operation: str = "operation") -> Any: ...

        def get_selected_entity_handles(self) -> List[str]: ...

        def _fast_get_property(self, obj: Any, property_name: str, default: Any = None) -> Any: ...

        def get_layers_info(self, entity_data: Any = None) -> List[Dict[str, Any]]: ...

        def resolve_export_path(self, filename: str, folder_type: str = "drawings") -> str: ...
        def list_blocks(self) -> List[str]: ...
        def get_block_counts(self, block_names: List[str] | None = None) -> Dict[str, int]: ...
        def get_block_info(self, block_name: str) -> Dict[str, Any]: ...

    def _get_entities_to_process(self, document: Any, only_selected: bool = False) -> list[Any]:
        """Get entities to process (all or selected).

        Optimized to use PickfirstSelectionSet for selected entities instead of
        iterating through entire ModelSpace.

        Args:
            document: AutoCAD document object
            only_selected: If True, get only selected entities. If False, get all.

        Returns:
            List of entity objects to process
        """
        entities_to_process = []

        if only_selected:
            # OPTIMIZED: Access selected entities directly from PickfirstSelectionSet
            # instead of iterating through entire ModelSpace looking for handles
            try:
                # Get the pickfirst selection set (current selection in AutoCAD)
                selection_set = document.PickfirstSelectionSet

                # Check if selection is empty
                if selection_set.Count == 0:
                    logger.info("No entities selected - returning empty list")
                    return []

                logger.info(f"Retrieving {selection_set.Count} selected entities")

                # Get entities directly from selection set (MUCH faster than iterating ModelSpace)
                for i in range(selection_set.Count):
                    try:
                        entities_to_process.append(selection_set.Item(i))
                    except Exception as e:
                        logger.debug(f"Failed to get selected entity at index {i}: {e}")
                        continue

            except Exception as e:
                logger.error(f"Failed to access PickfirstSelectionSet: {e}")
                logger.info("Falling back to handle-based selection method")

                # Fallback to old method if PickfirstSelectionSet fails
                selected_handles = self.get_selected_entity_handles()
                if not selected_handles:
                    logger.info("No entities selected - returning empty list")
                    return []

                logger.info(
                    f"Retrieving {len(selected_handles)} selected entities (fallback method)"
                )

                # Get entities by handle from ModelSpace
                try:
                    model_space = document.ModelSpace
                except Exception as e:
                    logger.error(f"Failed to access ModelSpace: {e}")
                    return []

                # Extract only selected entities
                selected_handles_set = set(selected_handles)

                for entity in model_space:
                    if str(entity.Handle) in selected_handles_set:
                        entities_to_process.append(entity)

        else:
            # Get all entities from ModelSpace
            try:
                model_space = document.ModelSpace
                entities_to_process = list(model_space)
            except Exception as e:
                logger.error(f"Failed to access ModelSpace: {e}")
                return []

        return entities_to_process

    def _extract_circle_properties(self, entity: Any) -> Dict[str, float]:
        """Extract Circle-specific geometry properties.

        Args:
            entity: Circle entity from AutoCAD

        Returns:
            Dictionary with radius, circumference, area, length
        """
        radius_val = self._fast_get_property(entity, "Radius")
        radius = float(radius_val) if radius_val else 0.0

        if radius > 0:
            circumference = 2 * math.pi * radius
            area = math.pi * radius * radius
        else:
            circumference = 0.0
            area = 0.0

        return {
            "Length": 0.0,
            "Area": round(area, 3) if area > 0 else 0.0,
            "Radius": round(radius, 3) if radius > 0 else 0.0,
            "Circumference": round(circumference, 3) if circumference > 0 else 0.0,
        }

    def _extract_arc_properties(self, entity: Any) -> Dict[str, float]:
        """Extract Arc-specific geometry properties.

        Args:
            entity: Arc entity from AutoCAD

        Returns:
            Dictionary with radius, length, circumference, area
        """
        radius_val = self._fast_get_property(entity, "Radius")
        length_val = self._fast_get_property(entity, "Length")

        radius = float(radius_val) if radius_val else 0.0
        length = float(length_val) if length_val else 0.0

        return {
            "Length": round(length, 3) if length > 0 else 0.0,
            "Area": 0.0,
            "Radius": round(radius, 3) if radius > 0 else 0.0,
            "Circumference": round(length, 3) if length > 0 else 0.0,  # Arc length
        }

    def _extract_line_properties(self, entity: Any) -> Dict[str, float]:
        """Extract Line-specific geometry properties.

        Args:
            entity: Line entity from AutoCAD

        Returns:
            Dictionary with length, area, radius, circumference
        """
        length_val = self._fast_get_property(entity, "Length")
        length = float(length_val) if length_val else 0.0

        return {
            "Length": round(length, 3) if length > 0 else 0.0,
            "Area": 0.0,
            "Radius": 0.0,
            "Circumference": 0.0,
        }

    def _extract_polyline_properties(self, entity: Any) -> Dict[str, float]:
        """Extract Polyline-specific geometry properties.

        Args:
            entity: Polyline entity from AutoCAD

        Returns:
            Dictionary with length, area, radius, circumference
        """
        length_val = self._fast_get_property(entity, "Length")
        area_val = self._fast_get_property(entity, "Area")

        length = float(length_val) if length_val else 0.0
        area = float(area_val) if area_val else 0.0

        return {
            "Length": round(length, 3) if length > 0 else 0.0,
            "Area": round(area, 3) if area > 0 else 0.0,
            "Radius": 0.0,
            "Circumference": 0.0,
        }

    def _extract_generic_properties(self, entity: Any) -> Dict[str, float]:
        """Extract generic entity properties (TEXT, DIMENSION, etc.).

        Args:
            entity: Generic entity from AutoCAD

        Returns:
            Dictionary with all geometry properties set to 0
        """
        return {
            "Length": 0.0,
            "Area": 0.0,
            "Radius": 0.0,
            "Circumference": 0.0,
        }

    def get_entity_counts(self) -> Dict[str, int]:
        """Get instant counts of main entity types using SelectionSets (O(K)).

        Returns:
            Dictionary with counts mapping internal name to count:
            { "Line": 120, "Polyline": 40, "LWPolyline": 15, "Circle":... }
        """
        import time

        try:
            self._validate_connection()
            document = self._get_document("get_entity_counts")

            type_counts = {}
            # Mapping of Friendly Name -> DXF Name for SelectionSet filtering
            # 0 = Object Type in DXF
            entity_type_map = {
                "Line": "LINE",
                "Polyline": "LWPOLYLINE",  # Modern polylines
                "Polyline2D": "POLYLINE",  # Old style/3D polylines
                "Circle": "CIRCLE",
                "Arc": "ARC",
                "Text": "TEXT",
                "MText": "MTEXT",
                "Block": "INSERT",
                "Spline": "SPLINE",
                "Ellipse": "ELLIPSE",
                "Hatch": "HATCH",
                "Dimension": "DIMENSION",
            }

            # Helper to convert to variant
            def to_variant_array(types, values):
                import win32com.client

                return win32com.client.VARIANT(types, values)

            import pythoncom

            ft = to_variant_array(pythoncom.VT_ARRAY | pythoncom.VT_I2, [0])

            perf_start = time.perf_counter()
            with self._selection_set_manager(document, "MCP_ENTITY_COUNTS") as ss:
                for clean_name, dxf_name in entity_type_map.items():
                    fd = to_variant_array(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, [dxf_name])
                    try:
                        ss.Clear()
                        ss.Select(5, None, None, ft, fd)  # 5 = acSelectionSetAll
                        count = ss.Count
                        if count > 0:
                            type_counts[clean_name] = count
                    except Exception as e:
                        logger.debug(f"Failed to count entity type {dxf_name}: {e}")
            elapsed = time.perf_counter() - perf_start
            logger.info(
                f"[PERF] Counted explicit entity types via SS in {elapsed:.3f}s: "
                f"{sum(type_counts.values())} total"
            )

            return type_counts

        except Exception as e:
            logger.error(f"Failed to get entity counts: {e}")
            return {}

    @contextmanager
    def _selection_set_manager(self, document: Any, name: str) -> Any:
        """Inline context manager for selection sets, to avoid circular imports."""
        try:
            document.SelectionSets.Item(name).Delete()
        except Exception:
            pass

        ss = document.SelectionSets.Add(name)
        try:
            yield ss
        finally:
            try:
                ss.Delete()
            except Exception:
                pass

    def _extract_single_entity_data(
        self,
        entity: Any,
        do_timing: bool,
        color_map_reverse: Dict[int, str],
        perf_property_times: Dict[str, float],
        com_call_stats: Dict[str, Any],
    ) -> Optional[Dict[str, Any]]:
        """Extract properties for a single entity, updating stats."""
        import math
        import time

        import win32com.client

        # ========== PHASE 1: Basic Properties (ALWAYS) ==========
        perf_t = time.perf_counter() if do_timing else 0.0

        # CRITICAL: Only fetch absolutely necessary properties
        handle = self._fast_get_property(entity, "Handle", "")
        object_type = self._fast_get_property(entity, "ObjectName", "Unknown")
        layer = self._fast_get_property(entity, "Layer", "0")
        calls_made = 3  # Handle, ObjectName, Layer

        if do_timing:
            perf_property_times["basic"] += time.perf_counter() - perf_t

        # Pre-process object type for fast lookups
        object_type_str = str(object_type)
        object_type_upper = object_type_str.upper()

        # OPTIMIZATION: Skip Color and Name for most entities (lazy fetch)
        needs_color = True  # Most entities need color
        needs_name = "BLOCK" in object_type_upper or "INSERT" in object_type_upper

        color = "ByLayer"
        name = ""

        if needs_color:
            # Use dynamic dispatch for more reliable property access
            dyn_entity = win32com.client.dynamic.Dispatch(entity)

            color_val = 256  # Default ByLayer
            try:
                # Standard Color property (ACI)
                raw_color = getattr(dyn_entity, "Color", 256)

                # If it's a TrueColor object, we need the ColorIndex
                if hasattr(raw_color, "ColorIndex"):
                    color_val = int(raw_color.ColorIndex)
                else:
                    color_val = int(raw_color)
            except (TypeError, ValueError, AttributeError):
                color_val = 256

            color = (
                "ByLayer" if color_val == 256 else color_map_reverse.get(color_val, str(color_val))
            )
            calls_made += 1
        else:
            com_call_stats["properties_skipped"] += 1

        if needs_name:
            name = self._fast_get_property(entity, "Name", "")
            calls_made += 1
        else:
            com_call_stats["properties_skipped"] += 1

        # ========== PHASE 2: Geometry Properties (SELECTIVE) ==========
        dyn_entity = win32com.client.dynamic.Dispatch(entity)

        length = 0.0
        area = 0.0
        radius = 0.0
        circumference = 0.0

        # OPTIMIZATION: Type-specific property extraction (minimize COM calls)
        if "CIRCLE" in object_type_upper:
            perf_t = time.perf_counter() if do_timing else 0.0
            radius_val = self._fast_get_property(dyn_entity, "Radius")
            calls_made += 1
            com_call_stats["calls_by_type"]["CIRCLE"] += 1
            com_call_stats["properties_skipped"] += 2

            if radius_val is not None:
                try:
                    radius = float(radius_val)
                    if radius > 0:
                        circumference = 2 * math.pi * radius
                        area = math.pi * radius * radius
                except (ValueError, TypeError):
                    pass
            if do_timing:
                perf_property_times["radius"] += time.perf_counter() - perf_t

        elif "ARC" in object_type_upper:
            perf_t = time.perf_counter() if do_timing else 0.0
            radius_val = self._fast_get_property(dyn_entity, "Radius")
            length_val = self._fast_get_property(dyn_entity, "Length")
            calls_made += 2
            com_call_stats["calls_by_type"]["ARC"] += 1
            com_call_stats["properties_skipped"] += 1

            if radius_val is not None:
                try:
                    radius = float(radius_val)
                except (ValueError, TypeError):
                    pass
            if length_val is not None:
                try:
                    length = float(length_val)
                    circumference = length  # Arc length
                except (ValueError, TypeError):
                    pass
            if do_timing:
                perf_property_times["radius"] += time.perf_counter() - perf_t

        elif (
            "LINE" in object_type_upper
            or "POLY" in object_type_upper
            or "SPLINE" in object_type_upper
        ):
            perf_t = time.perf_counter() if do_timing else 0.0
            length_val = self._fast_get_property(dyn_entity, "Length")
            calls_made += 1

            is_poly = "POLY" in object_type_upper
            is_spline = "SPLINE" in object_type_upper
            is_line = "LINE" in object_type_upper and not is_poly and not is_spline

            if is_line:
                com_call_stats["calls_by_type"]["LINE"] += 1
                com_call_stats["properties_skipped"] += 3
                if length_val is None or float(length_val) == 0:
                    try:
                        sp = self._fast_get_property(dyn_entity, "StartPoint")
                        ep = self._fast_get_property(dyn_entity, "EndPoint")
                        if sp is not None and ep is not None:
                            length_val = math.sqrt(sum((a - b) ** 2 for a, b in zip(sp, ep)))
                            calls_made += 2
                    except Exception:
                        pass
            elif is_poly:
                com_call_stats["calls_by_type"]["POLYLINE"] += 1
                if length_val is None or float(length_val) == 0:
                    alt_len = self._fast_get_property(dyn_entity, "TotalLength")
                    if alt_len is not None and float(alt_len) > 0:
                        length_val = alt_len
                        calls_made += 1

            if length_val is not None:
                try:
                    length = float(length_val)
                except (ValueError, TypeError):
                    pass

            if "POLY" in object_type_upper:
                area_val = self._fast_get_property(dyn_entity, "Area")
                calls_made += 1
                com_call_stats["properties_skipped"] += 2
                if area_val is not None:
                    try:
                        area = float(area_val)
                    except (ValueError, TypeError):
                        pass
            else:
                com_call_stats["properties_skipped"] += 1

            if do_timing:
                perf_property_times["geometry"] += time.perf_counter() - perf_t
        else:
            com_call_stats["calls_by_type"]["OTHER"] += 1
            com_call_stats["properties_skipped"] += 4

        com_call_stats["total_calls"] += calls_made

        return {
            "Handle": str(handle),
            "ObjectType": object_type_str,
            "Layer": str(layer).strip(),
            "Color": color,
            "Length": round(length, 3) if length > 0 else 0.0,
            "Area": round(area, 3) if area > 0 else 0.0,
            "Radius": round(radius, 3) if radius > 0 else 0.0,
            "Circumference": round(circumference, 3) if circumference > 0 else 0.0,
            "Name": str(name) if name else "",
        }

    def _populate_worksheet(
        self,
        worksheet: Any,
        columns: List[str],
        data_list: List[Dict[str, Any]],
        header_fill: Any,
        header_font: Any,
        center_cols: Optional[List[int]] = None,
        float_cols: Optional[List[int]] = None,
    ) -> None:
        """Helper to populate an Excel worksheet."""
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter

        # Write headers
        for col_idx, column_name in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            if cell is not None:
                cell.value = column_name
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data
        for row_idx, row_data in enumerate(data_list, 2):
            for col_idx, column_name in enumerate(columns, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell is not None:
                    value = row_data.get(column_name)
                    cell.value = value

                    if center_cols and col_idx in center_cols:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

                    if float_cols and col_idx in float_cols and isinstance(value, (int, float)):
                        cell.number_format = "0.000"

        # Auto-adjust column widths
        for col_idx, column_name in enumerate(columns, 1):
            max_length = len(column_name)
            for row_idx in range(2, len(data_list) + 2):
                cell_obj = worksheet.cell(row=row_idx, column=col_idx)
                cell_value = str(cell_obj.value or "") if cell_obj is not None else ""
                max_length = max(max_length, len(cell_value))
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

    def extract_drawing_data(
        self,
        only_selected: bool = False,
        limit: int = 500,
        offset: int = 0,
        entity_type: Optional[str] = None,
    ) -> list[dict]:
        """Extract drawing data (entities) with their properties.

        Optimized iteration through ModelSpace or selected entities with reduced COM calls.
        Uses property caching and batch processing for improved performance.

        Args:
            only_selected: If True, extract only selected entities. If False, extract all.
                          Defaults to False for backward compatibility.
            limit: Maximum number of entities to return per page. Defaults to 500.
            offset: Number of entities to skip before extracting. Defaults to 0.
            entity_type: Optional DXF type name to filter by (e.g., 'LINE', 'LWPOLYLINE').

        Returns:
            List of dictionaries with columns:
            - Handle: Entity handle (unique identifier)
            - ObjectType: Type of object (LINE, CIRCLE, LWPOLYLINE, etc.)
            - Layer: Layer name
            - Color: Color index (0-255) or color name
            - Length: Length (for linear objects)
            - Area: Area (for closed objects)
            - Radius: Radius (for circles and arcs)
            - Circumference: Circumference (2πr for circles, arc length for arcs)
            - Name: Name (for blocks, layers, etc.)
        """
        import time

        perf_start_total = time.perf_counter()

        try:
            self._validate_connection()
            document = self._get_document("extract_drawing_data")
            entities_data = []

            # Optimized selection: If we have an entity_type and we are NOT in selection-only mode,
            # use a dedicated SelectionSet filter for maximum performance.
            perf_start_selection = time.perf_counter()
            items_iterator = None
            total_available = 0

            if entity_type and not only_selected:
                # FAST PATH: Use SelectionSet filter for specific type
                import pythoncom

                def to_variant_array(types, values):
                    import win32com.client

                    return win32com.client.VARIANT(types, values)

                ft = to_variant_array(
                    pythoncom.VT_ARRAY | pythoncom.VT_I2, [0]
                )  # DXF Code 0 (Type)
                fd = to_variant_array(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, [entity_type])

                # Use a specific SS name for this extraction
                ss_name = f"MCP_EXTRACT_{entity_type}"
                with self._selection_set_manager(document, ss_name) as ss:
                    ss.Select(5, None, None, ft, fd)  # 5 = acSelectionSetAll
                    total_available = ss.Count

                    # Manual slicing on the selection set is very fast
                    start_idx = min(offset, total_available)
                    end_idx = min(offset + limit, total_available)

                    extracted_items = []
                    for i in range(start_idx, end_idx):
                        try:
                            extracted_items.append(ss.Item(i))
                        except Exception:
                            continue
                    items_iterator = extracted_items
            else:
                # SLOW PATH: Use existing selection logic
                entities_to_process = self._get_entities_to_process(document, only_selected)
                total_available = len(entities_to_process)
                import itertools

                items_iterator = (
                    list(itertools.islice(entities_to_process, offset, offset + limit))
                    if limit > 0
                    else entities_to_process
                )

            perf_selection_time = time.perf_counter() - perf_start_selection

            if not items_iterator:
                logger.info("No entities to process - returning empty data")
                return []

            logger.info(
                f"[PERF] Entity selection ({entity_type or 'ALL'}) took "
                f"{perf_selection_time:.3f}s. "
                f"Extracting range {offset}-{offset + limit} of {total_available}"
            )

            # Pre-build reverse color map for faster lookups
            color_map_reverse = {v: k for k, v in COLOR_MAP.items()}
            import itertools

            # Optimized iteration with reduced COM calls
            entity_count = 0
            error_count = 0
            perf_start_iteration = time.perf_counter()

            # Timing stats for property extraction
            perf_property_times: Dict[str, float] = {
                "basic": 0.0,  # Handle, ObjectName, Layer, Color, Name
                "geometry": 0.0,  # Length, Area
                "radius": 0.0,  # Radius, Circumference
            }

            # COM call statistics
            com_call_stats: Dict[str, Any] = {
                "total_calls": 0,
                "calls_by_type": {
                    "CIRCLE": 0,
                    "ARC": 0,
                    "LINE": 0,
                    "POLYLINE": 0,
                    "OTHER": 0,
                },
                "properties_skipped": 0,
            }
            # Progress tracking for large datasets
            progress_interval = 1000  # Log every 1000 entities
            sample_interval = 100  # Sample detailed timing every N entities

            for entity in items_iterator:
                entity_count += 1

                # Progress logging for large datasets
                if entity_count % progress_interval == 0:
                    elapsed = time.perf_counter() - perf_start_iteration
                    rate = entity_count / elapsed if elapsed > 0 else 0
                    logger.info(
                        f"[PERF] Progress: {entity_count} entities extracted "
                        f"- {rate:.1f} entities/s"
                    )

                try:
                    # Sample timing on a subset of entities
                    do_timing = entity_count % sample_interval == 0

                    entity_data = self._extract_single_entity_data(
                        entity,
                        do_timing,
                        color_map_reverse,
                        perf_property_times,
                        com_call_stats,
                    )
                    if entity_data:
                        entities_data.append(entity_data)

                except Exception as e:
                    logger.debug(f"Failed to extract entity data (entity #{entity_count}): {e}")
                    error_count += 1
                    continue

            perf_iteration_time = time.perf_counter() - perf_start_iteration
            perf_total_time = time.perf_counter() - perf_start_total

            logger.info(
                f"Extracted data from {len(entities_data)} entities "
                f"(processed {entity_count}, {error_count} errors)"
            )
            logger.info(
                f"[PERF] Entity iteration/extraction took {perf_iteration_time:.3f}s "
                f"({entity_count / perf_iteration_time:.1f} entities/s)"
            )

            # Detailed property timing breakdown
            samples_count = entity_count // sample_interval
            if samples_count > 0:
                avg_basic = (perf_property_times["basic"] / samples_count) * 1000
                avg_geometry = (perf_property_times["geometry"] / samples_count) * 1000
                avg_radius = (perf_property_times["radius"] / samples_count) * 1000
                logger.info(
                    f"[PERF] Property extraction (avg per entity): "
                    f"basic={avg_basic:.2f}ms, geometry={avg_geometry:.2f}ms, "
                    f"radius={avg_radius:.2f}ms"
                )

            # COM call optimization statistics
            total_calls = com_call_stats["total_calls"]
            skipped_calls = com_call_stats["properties_skipped"]
            potential_calls = total_calls + skipped_calls
            savings_pct = (skipped_calls / potential_calls * 100) if potential_calls > 0 else 0

            logger.info(
                f"[PERF] COM calls: {total_calls:,} made, {skipped_calls:,} skipped "
                f"({savings_pct:.1f}% reduction)"
            )
            logger.info(
                f"[PERF] Entity type breakdown: "
                f"CIRCLE={com_call_stats['calls_by_type']['CIRCLE']}, "
                f"ARC={com_call_stats['calls_by_type']['ARC']}, "
                f"LINE={com_call_stats['calls_by_type']['LINE']}, "
                f"POLY={com_call_stats['calls_by_type']['POLYLINE']}, "
                f"OTHER={com_call_stats['calls_by_type']['OTHER']}"
            )

            logger.info(f"[PERF] Total extraction time: {perf_total_time:.3f}s")
            return entities_data

        except Exception as e:
            logger.error(f"Failed to extract drawing data: {e}")
            return []

    def export_to_excel(self, filepath: str = "drawing_data.xlsx") -> bool:
        """Export drawing data to Excel file.

        Uses the configured output directory from config.json for security,
        similar to save_drawing(). If only filename provided, saves to output directory.

        Args:
            filepath: Path to output Excel file (default: "drawing_data.xlsx")
                     - If filename only, saved to config output directory
                     - If path provided, must be within output directory

        Returns:
            True if successful, False otherwise
        """
        import time

        perf_start_total = time.perf_counter()

        try:
            from pathlib import Path

            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            perf_start_setup = time.perf_counter()

            # Resolve final path using centralized utility
            if not filepath or filepath == "drawing_data.xlsx":
                try:
                    document = self._get_document("export_to_excel")
                    doc_stem = Path(document.Name).stem if document and document.Name else "drawing"
                    filename = f"{doc_stem}_data.xlsx"
                except Exception:
                    filename = "drawing_data.xlsx"
            else:
                filename = Path(filepath).name
            full_filepath_str = self.resolve_export_path(filename, "sheets")
            full_filepath = Path(full_filepath_str)

            perf_setup_time = time.perf_counter() - perf_start_setup
            logger.info(f"[PERF] Export setup took {perf_setup_time:.3f}s")

            # Extract data
            perf_start_extract = time.perf_counter()
            data = self.extract_drawing_data(limit=0)
            perf_extract_time = time.perf_counter() - perf_start_extract
            logger.info(f"[PERF] Data extraction took {perf_extract_time:.3f}s")

            if not data:
                logger.warning("No data to export")
                return False

            # Create workbook
            perf_start_workbook = time.perf_counter()
            workbook: Any = Workbook()
            worksheet: Any = workbook.active
            if worksheet is None:
                logger.error("Failed to create worksheet")
                return False

            worksheet.title = "Drawing Data"

            # Define columns
            columns = [
                "Handle",
                "ObjectType",
                "Layer",
                "Color",
                "Length",
                "Area",
                "Radius",
                "Circumference",
                "Name",
            ]

            # Write headers with styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            self._populate_worksheet(
                worksheet=worksheet,
                columns=columns,
                data_list=data,
                header_fill=header_fill,
                header_font=header_font,
                float_cols=[5, 6, 7, 8],  # Length, Area, Radius, Circumference
            )

            perf_workbook_time = time.perf_counter() - perf_start_workbook
            logger.info(f"[PERF] Workbook population took {perf_workbook_time:.3f}s")

            # ========== Create Layers Sheet ==========
            perf_start_layers = time.perf_counter()
            # OPTIMIZED: Pass extracted data to avoid re-iterating ModelSpace
            layers_info = self.get_layers_info(entity_data=data)
            layers_sheet: Any = workbook.create_sheet("Layers")
            if layers_sheet is not None:
                # Define columns for layers sheet
                layer_columns = [
                    "Name",
                    "ObjectCount",
                    "Color",
                    "Linetype",
                    "Lineweight",
                    "IsLocked",
                    "IsVisible",
                ]

                self._populate_worksheet(
                    worksheet=layers_sheet,
                    columns=layer_columns,
                    data_list=layers_info,
                    header_fill=header_fill,
                    header_font=header_font,
                    center_cols=[2, 6, 7],  # ObjectCount, IsLocked, IsVisible
                )

            perf_layers_time = time.perf_counter() - perf_start_layers
            logger.info(f"[PERF] Creating layers sheet took {perf_layers_time:.3f}s")

            # ========== Create Blocks Sheet ==========
            perf_start_blocks = time.perf_counter()
            block_names = self.list_blocks()
            insert_counts = self.get_block_counts(block_names)
            blocks_sheet: Any = workbook.create_sheet("Blocks")

            if blocks_sheet is not None:
                # Define columns for blocks sheet
                block_columns = [
                    "Name",
                    "InstanceCount",
                    "ObjectCount",
                    "IsXRef",
                    "Comments",
                ]

                # Prepare block data
                block_data_list = []
                for bname in block_names:
                    info = self.get_block_info(bname)
                    block_data_list.append(
                        {
                            "Name": bname,
                            "InstanceCount": insert_counts.get(bname, 0),
                            "ObjectCount": info.get("ObjectCount", 0),
                            "IsXRef": info.get("IsXRef", False),
                            "Comments": info.get("Comments", ""),
                        }
                    )

                self._populate_worksheet(
                    worksheet=blocks_sheet,
                    columns=block_columns,
                    data_list=block_data_list,
                    header_fill=header_fill,
                    header_font=header_font,
                    center_cols=[2, 3, 4],  # InstanceCount, ObjectCount, IsXRef
                )
            perf_blocks_time = time.perf_counter() - perf_start_blocks
            logger.info(f"[PERF] Creating blocks sheet took {perf_blocks_time:.3f}s")

            # Save workbook
            perf_start_save = time.perf_counter()
            workbook.save(str(full_filepath))
            perf_save_time = time.perf_counter() - perf_start_save
            perf_total_time = time.perf_counter() - perf_start_total

            logger.info(f"[PERF] Saving workbook took {perf_save_time:.3f}s")
            logger.info(f"[PERF] Total export time: {perf_total_time:.3f}s")
            logger.info(
                f"Exported {len(data)} entities, {len(layers_info)} layers, and "
                f"{len(block_names)} blocks to {full_filepath}"
            )
            return True

        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            return False
        except Exception as e:
            logger.error(f"Failed to export to Excel: {e}")
            return False
