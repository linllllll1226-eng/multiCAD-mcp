"""Unit tests for CAD adapters.

Test adapter interface implementation and context managers.
"""

from unittest.mock import MagicMock, PropertyMock, patch

import pytest

from src.adapters import AutoCADAdapter
from src.core import CADInterface


class TestAdapterInstantiation:
    """Test suite for adapter instantiation."""

    def test_adapter_can_be_instantiated(self):
        """Test that AutoCADAdapter can be instantiated."""
        adapter = AutoCADAdapter("autocad")
        assert adapter is not None
        assert adapter.cad_type == "autocad"

    def test_adapter_supports_multiple_cad_types(self):
        """Test that AutoCADAdapter supports multiple CAD types."""
        # All should work with AutoCADAdapter
        cad_types = ["autocad", "zwcad", "gcad", "bricscad"]
        for cad_type in cad_types:
            adapter = AutoCADAdapter(cad_type)
            assert adapter.cad_type == cad_type

    def test_adapter_case_insensitive(self):
        """Test that cad_type is case-insensitive."""
        adapter1 = AutoCADAdapter("autocad")
        adapter2 = AutoCADAdapter("AUTOCAD")
        assert adapter1.cad_type == adapter2.cad_type


class TestContextManagers:
    """Test suite for context managers."""

    def test_adapter_context_manager(self):
        """Test that AutoCADAdapter can be used as context manager."""
        adapter = AutoCADAdapter("autocad")

        with (
            patch.object(adapter, "connect", return_value=True) as mock_connect,
            patch.object(adapter, "disconnect") as mock_disconnect,
        ):
            # Use adapter as context manager
            with adapter as ctx_adapter:
                assert ctx_adapter is adapter
                mock_connect.assert_called_once()

            # Should disconnect on exit
            mock_disconnect.assert_called_once()

    def test_adapter_context_manager_raises_on_connection_failure(self):
        """Test that context manager raises if connection fails."""
        adapter = AutoCADAdapter("autocad")

        with patch.object(adapter, "connect", return_value=False):
            # Should raise CADConnectionError if connection fails
            from core.exceptions import CADConnectionError

            with pytest.raises(CADConnectionError):
                with adapter:
                    pass

    def test_com_session_context_manager(self):
        """Test that com_session properly initializes/uninitializes COM."""
        import sys

        from src.adapters.autocad_adapter import com_session

        if sys.platform != "win32":
            pytest.skip("COM only available on Windows")

        # Import pythoncom for mocking
        import pythoncom

        with (
            patch.object(pythoncom, "CoInitialize") as mock_init,
            patch.object(pythoncom, "CoUninitialize") as mock_uninit,
        ):
            with com_session():
                mock_init.assert_called_once()

            # Should uninitialize on exit
            mock_uninit.assert_called_once()

    def test_selection_set_manager_creates_and_deletes(self):
        """Test that SelectionSetManager creates and deletes selection sets."""
        from src.adapters.autocad_adapter import SelectionSetManager

        mock_document = MagicMock()
        mock_ss = MagicMock()
        mock_document.SelectionSets.Add.return_value = mock_ss

        with SelectionSetManager(mock_document, "TEST_SS") as ss:
            # Should create selection set
            mock_document.SelectionSets.Add.assert_called_once_with("TEST_SS")
            assert ss is mock_ss

        # Should delete on exit
        mock_ss.Delete.assert_called_once()

    def test_selection_set_manager_handles_existing(self):
        """Test that SelectionSetManager deletes existing selection set."""
        from src.adapters.autocad_adapter import SelectionSetManager

        mock_document = MagicMock()
        mock_existing_ss = MagicMock()
        mock_new_ss = MagicMock()

        # Simulate existing selection set
        mock_document.SelectionSets.Item.return_value = mock_existing_ss
        mock_document.SelectionSets.Add.return_value = mock_new_ss

        with SelectionSetManager(mock_document, "TEST_SS") as ss:
            # Should delete existing
            mock_existing_ss.Delete.assert_called_once()
            # Should create new
            mock_document.SelectionSets.Add.assert_called_once_with("TEST_SS")
            assert ss is mock_new_ss


class TestCADInterfaceContract:
    """Test suite for CADInterface contract compliance."""

    def test_adapter_implements_cad_interface(self):
        """Test that AutoCADAdapter implements CADInterface."""
        from src.adapters.autocad_adapter import AutoCADAdapter

        # Verify adapter inherits from CADInterface (check in MRO)
        mro_names = [c.__name__ for c in AutoCADAdapter.__mro__]
        assert "CADInterface" in mro_names, f"CADInterface must be in MRO. Found: {mro_names}"

        # Also check that adapter has the ABC.ABCMeta metaclass (from CADInterface)
        assert hasattr(AutoCADAdapter, "__abstractmethods__"), (
            "AutoCADAdapter should have __abstractmethods__ from ABC"
        )

    def test_adapter_has_required_methods(self):
        """Test that adapter has all required interface methods."""
        required_methods = [
            # Connection
            "connect",
            "disconnect",
            "is_connected",
            # Drawing
            "draw_line",
            "draw_circle",
            "draw_arc",
            "draw_rectangle",
            "draw_polyline",
            "draw_ellipse",
            "draw_text",
            "draw_hatch",
            "add_dimension",
            # Layers
            "create_layer",
            "set_current_layer",
            "get_current_layer",
            "list_layers",
            # Files
            "save_drawing",
            "open_drawing",
            "new_drawing",
            # View
            "zoom_extents",
            "refresh_view",
            # Entity
            "delete_entity",
            "get_entity_properties",
            "set_entity_properties",
            # Data extraction
            "extract_drawing_data",
            "get_layers_info",
        ]

        adapter = AutoCADAdapter("autocad")
        for method in required_methods:
            assert hasattr(adapter, method), f"AutoCADAdapter missing {method}"


class TestCoordinateNormalization:
    """Test suite for coordinate normalization."""

    def test_normalize_2d_coordinate(self):
        """Test normalization of 2D coordinates."""
        result = CADInterface.normalize_coordinate((10, 20))
        assert result == (10.0, 20.0, 0.0)

    def test_normalize_3d_coordinate(self):
        """Test normalization of 3D coordinates."""
        result = CADInterface.normalize_coordinate((10, 20, 30))
        assert result == (10.0, 20.0, 30.0)

    def test_normalize_coordinate_converts_to_float(self):
        """Test that normalize_coordinate converts to float."""
        result = CADInterface.normalize_coordinate((10, 20))
        assert isinstance(result[0], float)
        assert isinstance(result[1], float)
        assert isinstance(result[2], float)

    def test_normalize_coordinate_invalid_raises(self):
        """Test invalid coordinate raises error."""
        with pytest.raises(ValueError):
            # Only 1 dimension - type: ignore
            CADInterface.normalize_coordinate((10,))  # type: ignore

        with pytest.raises(ValueError):
            # Too many dimensions - type: ignore
            CADInterface.normalize_coordinate((10, 20, 30, 40))  # type: ignore


class TestLineWeightValidation:
    """Test suite for lineweight validation."""

    def test_valid_lineweight_accepted(self):
        """Test that valid lineweights are accepted."""
        adapter = AutoCADAdapter("autocad")
        valid_weights = [0, 5, 9, 13, 15, 18, 20, 25, 30, 100, 211]
        for weight in valid_weights:
            assert adapter.validate_lineweight(weight) == weight

    def test_invalid_lineweight_returns_default(self):
        """Test that invalid lineweights return default."""
        adapter = AutoCADAdapter("autocad")
        result = adapter.validate_lineweight(999)
        assert result == 0  # Default thin line

    def test_lineweight_is_valid_check(self):
        """Test lineweight validity check."""
        from src.core.cad_interface import LineWeight

        assert LineWeight.is_valid(0) is True
        assert LineWeight.is_valid(50) is True
        assert LineWeight.is_valid(211) is True
        assert LineWeight.is_valid(999) is False
        assert LineWeight.is_valid(-1) is False


class TestColorEnums:
    """Test suite for color enumerations."""

    def test_color_enum_has_standard_colors(self):
        """Test that Color enum has standard colors."""
        from src.core.cad_interface import Color

        assert hasattr(Color, "BLACK")
        assert hasattr(Color, "RED")
        assert hasattr(Color, "GREEN")
        assert hasattr(Color, "BLUE")
        assert hasattr(Color, "WHITE")

    def test_color_enum_values_are_rgb_tuples(self):
        """Test that color values are RGB tuples."""
        from src.core.cad_interface import Color

        for color in Color:
            assert isinstance(color.value, tuple)
            assert len(color.value) == 3
            # Check RGB values are in valid range
            for component in color.value:
                assert 0 <= component <= 255


class TestRefreshViewUndoRedo:
    """Test suite for refresh_view and undo/redo interaction."""

    def test_refresh_view_uses_multiple_techniques(self):
        """Test that refresh_view uses multiple fallback techniques.

        Techniques in order:
        1. application.Refresh() (COM API - no undo/redo impact)
        2. SendCommand with REDRAW (most reliable visual update)
        3. Window click simulation (forces UI update)
        """
        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter = AutoCADAdapter("autocad")

        with (
            patch.object(adapter, "_get_application") as mock_get_app,
            patch.object(adapter, "_get_document") as mock_get_doc,
            patch.object(adapter, "_simulate_autocad_click") as mock_click,
        ):
            mock_app = MagicMock()
            mock_doc = MagicMock()
            mock_get_app.return_value = mock_app
            mock_get_doc.return_value = mock_doc

            result = adapter.refresh_view()

            assert result is True
            # Technique 1: COM API Refresh (should be tried first)
            mock_app.Refresh.assert_called_once()
            # Technique 2: REDRAW command
            mock_doc.SendCommand.assert_called_once_with("_redraw\n")
            # Technique 3: Window click simulation
            mock_click.assert_called_once()

    def test_undo_does_not_call_refresh_view(self):
        """Test that undo() does not call refresh_view().

        Removing refresh_view from undo/redo avoids contaminating
        the undo/redo stack with extra refresh commands.
        """
        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter = AutoCADAdapter("autocad")

        with (
            patch.object(adapter, "_validate_connection"),
            patch.object(adapter, "_get_application") as mock_get_app,
            patch.object(adapter, "refresh_view") as mock_refresh,
        ):
            mock_app = MagicMock()
            mock_get_app.return_value = mock_app

            # Call undo
            result = adapter.undo(count=1)

            assert result is True
            # Verify refresh_view was NOT called
            mock_refresh.assert_not_called()
            # Verify undo command was sent
            mock_app.ActiveDocument.SendCommand.assert_called_once_with("_undo 1\n")

    def test_redo_does_not_call_refresh_view(self):
        """Test that redo() does not call refresh_view().

        Removing refresh_view from undo/redo avoids contaminating
        the undo/redo stack with extra refresh commands.
        """
        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter = AutoCADAdapter("autocad")

        with (
            patch.object(adapter, "_validate_connection"),
            patch.object(adapter, "_get_application") as mock_get_app,
            patch.object(adapter, "refresh_view") as mock_refresh,
        ):
            mock_app = MagicMock()
            mock_get_app.return_value = mock_app

            # Call redo
            result = adapter.redo(count=1)

            assert result is True
            # Verify refresh_view was NOT called
            mock_refresh.assert_not_called()
            # Verify redo command was sent
            mock_app.ActiveDocument.SendCommand.assert_called_once_with("_redo 1\n")


class TestDataExport:
    """Test suite for drawing data extraction and export."""

    def test_extract_drawing_data_returns_list(self):
        """Test that extract_drawing_data returns a list of dictionaries."""
        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter = AutoCADAdapter("autocad")

        with (
            patch.object(adapter, "_validate_connection"),
            patch.object(adapter, "_get_document") as mock_get_doc,
            patch.object(adapter, "_get_entities_to_process") as mock_get_entities,
        ):
            # Mock entity with properties
            mock_entity = MagicMock()
            mock_entity.Handle = "A1B2C3D4"
            mock_entity.ObjectName = "AcDbLine"
            mock_entity.Layer = "0"
            mock_entity.Color = 1
            mock_entity.Length = 100.0
            mock_entity.Area = 0.0
            mock_entity.Radius = 0.0
            mock_entity.Name = ""

            # Mock _get_entities_to_process to return list of entities
            mock_get_entities.return_value = [mock_entity]

            mock_doc = MagicMock()
            mock_get_doc.return_value = mock_doc

            result = adapter.extract_drawing_data()

            assert isinstance(result, list)
            assert len(result) == 1
            assert result[0]["Handle"] == "A1B2C3D4"
            assert result[0]["ObjectType"] == "AcDbLine"
            assert result[0]["Layer"] == "0"
            # Verify _get_entities_to_process was called with only_selected=False
            mock_get_entities.assert_called_once_with(mock_doc, False)

    def test_extract_drawing_data_handles_empty_drawing(self):
        """Test that extract_drawing_data handles empty drawings."""
        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter = AutoCADAdapter("autocad")

        with (
            patch.object(adapter, "_validate_connection"),
            patch.object(adapter, "_get_document") as mock_get_doc,
            patch.object(adapter, "_get_entities_to_process") as mock_get_entities,
        ):
            # Mock empty entity list
            mock_get_entities.return_value = []

            mock_doc = MagicMock()
            mock_get_doc.return_value = mock_doc

            result = adapter.extract_drawing_data()

            assert isinstance(result, list)
            assert len(result) == 0

    def test_extract_drawing_data_with_only_selected(self):
        """Test that extract_drawing_data respects only_selected parameter."""
        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter = AutoCADAdapter("autocad")

        with (
            patch.object(adapter, "_validate_connection"),
            patch.object(adapter, "_get_document") as mock_get_doc,
            patch.object(adapter, "_get_entities_to_process") as mock_get_entities,
        ):
            # Mock selected entity
            mock_entity = MagicMock()
            mock_entity.Handle = "SELECTED1"
            mock_entity.ObjectName = "AcDbCircle"
            mock_entity.Layer = "1"
            mock_entity.Color = 5
            mock_entity.Radius = 50.0
            mock_entity.Length = 0.0
            mock_entity.Area = 0.0
            mock_entity.Name = ""

            mock_get_entities.return_value = [mock_entity]

            mock_doc = MagicMock()
            mock_get_doc.return_value = mock_doc

            # Call with only_selected=True
            result = adapter.extract_drawing_data(only_selected=True)

            assert isinstance(result, list)
            assert len(result) == 1
            assert result[0]["Handle"] == "SELECTED1"
            # Verify _get_entities_to_process was called with only_selected=True
            mock_get_entities.assert_called_once_with(mock_doc, True)

    def test_get_entities_to_process_uses_pickfirst_selection(self):
        """Test that _get_entities_to_process uses PickfirstSelectionSet for selected entities."""
        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter = AutoCADAdapter("autocad")

        # Mock document with PickfirstSelectionSet
        mock_doc = MagicMock()
        mock_selection = MagicMock()
        mock_selection.Count = 2

        # Mock selected entities
        mock_entity1 = MagicMock()
        mock_entity1.Handle = "E1"
        mock_entity2 = MagicMock()
        mock_entity2.Handle = "E2"

        mock_selection.Item.side_effect = [mock_entity1, mock_entity2]
        mock_doc.PickfirstSelectionSet = mock_selection

        # Call _get_entities_to_process with only_selected=True
        result = adapter._get_entities_to_process(mock_doc, only_selected=True)

        assert len(result) == 2
        assert result[0] is mock_entity1
        assert result[1] is mock_entity2
        # Verify PickfirstSelectionSet was accessed
        assert mock_selection.Item.call_count == 2

    def test_get_entities_to_process_returns_all_from_modelspace(self):
        """Test that _get_entities_to_process returns all entities from ModelSpace."""
        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter = AutoCADAdapter("autocad")

        # Mock document with ModelSpace
        mock_doc = MagicMock()
        mock_model_space = MagicMock()

        # Mock entities in ModelSpace
        mock_entity1 = MagicMock()
        mock_entity1.Handle = "E1"
        mock_entity2 = MagicMock()
        mock_entity2.Handle = "E2"
        mock_entity3 = MagicMock()
        mock_entity3.Handle = "E3"

        # Make ModelSpace iterable
        mock_model_space.__iter__.return_value = iter([mock_entity1, mock_entity2, mock_entity3])
        mock_doc.ModelSpace = mock_model_space

        # Call _get_entities_to_process with only_selected=False
        result = adapter._get_entities_to_process(mock_doc, only_selected=False)

        assert len(result) == 3
        assert result[0] is mock_entity1
        assert result[1] is mock_entity2
        assert result[2] is mock_entity3

    def test_export_to_excel_handles_no_data(self):
        """Test that export_to_excel handles drawing with no data."""
        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter = AutoCADAdapter("autocad")

        with patch.object(adapter, "extract_drawing_data") as mock_extract:
            # Mock empty drawing data
            mock_extract.return_value = []

            result = adapter.export_to_excel("/tmp/test.xlsx")

            # Should return False when there's no data
            assert result is False

    def test_export_to_excel_creates_file(self, tmp_path, monkeypatch):
        """Test that export_to_excel creates a valid Excel file."""
        monkeypatch.setenv("MULTICAD_OUTPUT_DIR", str(tmp_path / "导出"))
        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter = AutoCADAdapter("autocad")

        with (
            patch.object(adapter, "extract_drawing_data") as mock_extract,
            patch.object(adapter, "get_layers_info") as mock_get_layers_info,
        ):
            # Mock drawing data
            mock_extract.return_value = [
                {
                    "Handle": "A1B2C3D4",
                    "ObjectType": "AcDbLine",
                    "Layer": "0",
                    "Color": "red",
                    "Length": "100.500",
                    "Area": "0.000",
                    "Name": "",
                },
                {
                    "Handle": "E5F6G7H8",
                    "ObjectType": "AcDbCircle",
                    "Layer": "1",
                    "Color": "blue",
                    "Length": "314.159",
                    "Area": "7853.981",
                    "Name": "",
                },
            ]

            # Mock layers info
            mock_get_layers_info.return_value = [
                {
                    "Name": "0",
                    "ObjectCount": 1,
                    "Color": "white",
                    "Linetype": "Continuous",
                    "Lineweight": "Default",
                    "IsLocked": False,
                    "IsVisible": True,
                },
                {
                    "Name": "1",
                    "ObjectCount": 0,
                    "Color": "red",
                    "Linetype": "Continuous",
                    "Lineweight": "Default",
                    "IsLocked": False,
                    "IsVisible": True,
                },
                {
                    "Name": "MyLayer",
                    "ObjectCount": 2,
                    "Color": "blue",
                    "Linetype": "Dashed",
                    "Lineweight": "0.5",
                    "IsLocked": True,
                    "IsVisible": False,
                },
            ]

            # Use filename that will be saved to configured output directory
            filename = "test_export.xlsx"

            try:
                result = adapter.export_to_excel(filename)

                # Should succeed if openpyxl is installed
                assert result is True

                # Get expected filepath
                output_dir = (tmp_path / "导出").resolve()
                filepath = output_dir / "sheets" / filename

                # File should exist
                assert filepath.exists()
                # File should have content
                assert filepath.stat().st_size > 0
            finally:
                # Cleanup
                output_dir = (tmp_path / "导出").resolve()
                filepath = output_dir / "sheets" / filename
                if filepath.exists():
                    filepath.unlink()

    def test_export_to_excel_creates_layers_sheet(self, tmp_path, monkeypatch):
        """Test that export_to_excel creates a Layers sheet with detailed layer information."""
        monkeypatch.setenv("MULTICAD_OUTPUT_DIR", str(tmp_path / "图层导出"))
        from openpyxl import load_workbook

        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter = AutoCADAdapter("autocad")

        with (
            patch.object(adapter, "extract_drawing_data") as mock_extract,
            patch.object(adapter, "get_layers_info") as mock_get_layers_info,
        ):
            # Mock drawing data
            mock_extract.return_value = [
                {
                    "Handle": "A1B2C3D4",
                    "ObjectType": "AcDbLine",
                    "Layer": "0",
                    "Color": "red",
                    "Length": "100.500",
                    "Area": "0.000",
                    "Name": "",
                },
            ]

            # Mock layers info with detailed data
            mock_get_layers_info.return_value = [
                {
                    "Name": "0",
                    "ObjectCount": 1,
                    "Color": "white",
                    "Linetype": "Continuous",
                    "Lineweight": "Default",
                    "IsLocked": False,
                    "IsVisible": True,
                },
                {
                    "Name": "1",
                    "ObjectCount": 0,
                    "Color": "red",
                    "Linetype": "Continuous",
                    "Lineweight": "Default",
                    "IsLocked": False,
                    "IsVisible": True,
                },
                {
                    "Name": "MyLayer",
                    "ObjectCount": 2,
                    "Color": "blue",
                    "Linetype": "Dashed",
                    "Lineweight": "0.5",
                    "IsLocked": True,
                    "IsVisible": False,
                },
            ]

            filename = "test_layers.xlsx"

            try:
                result = adapter.export_to_excel(filename)
                assert result is True

                # Get expected filepath
                output_dir = (tmp_path / "图层导出").resolve()
                filepath = output_dir / "sheets" / filename

                # Load the workbook and check sheets
                workbook = load_workbook(str(filepath))
                sheet_names = workbook.sheetnames

                # Should have both Drawing Data and Layers sheets
                assert "Drawing Data" in sheet_names
                assert "Layers" in sheet_names

                # Check Layers sheet content
                layers_sheet = workbook["Layers"]
                # Header + 3 layers = 4 rows
                assert layers_sheet.max_row >= 4
                # Check headers
                assert layers_sheet.cell(row=1, column=1).value == "Name"
                assert layers_sheet.cell(row=1, column=2).value == "ObjectCount"
                assert layers_sheet.cell(row=1, column=3).value == "Color"
                # Check layer data (Name column)
                assert layers_sheet.cell(row=2, column=1).value == "0"
                assert layers_sheet.cell(row=3, column=1).value == "1"
                assert layers_sheet.cell(row=4, column=1).value == "MyLayer"
                # Check object counts
                assert layers_sheet.cell(row=2, column=2).value == 1
                assert layers_sheet.cell(row=3, column=2).value == 0
                assert layers_sheet.cell(row=4, column=2).value == 2

            finally:
                # Cleanup
                output_dir = (tmp_path / "图层导出").resolve()
                filepath = output_dir / "sheets" / filename
                if filepath.exists():
                    filepath.unlink()

    def test_extract_drawing_data_structure(self):
        """Test that extract_drawing_data returns proper dict structure."""
        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter = AutoCADAdapter("autocad")

        # extract_drawing_data returns list of dicts with entity data
        # We test that it returns empty list when disconnected (can't access real drawing)
        result = adapter.extract_drawing_data()

        assert isinstance(result, list)
        # When not connected to AutoCAD, should return empty list
        # Each item should have the required keys
        if result:
            for entity_dict in result:
                assert isinstance(entity_dict, dict)
                assert "Handle" in entity_dict
                assert "ObjectType" in entity_dict
                assert "Layer" in entity_dict
                assert "Color" in entity_dict
                assert "Length" in entity_dict
                assert "Area" in entity_dict
                assert "Radius" in entity_dict
                assert "Circumference" in entity_dict
                assert "Name" in entity_dict

    def test_color_map_reverse_in_extract_data(self):
        """Test that color mapping works in extract_drawing_data."""
        from src.adapters.autocad_adapter import COLOR_MAP, AutoCADAdapter

        AutoCADAdapter("autocad")

        # Verify COLOR_MAP exists and has expected values
        assert isinstance(COLOR_MAP, dict)
        assert "red" in COLOR_MAP
        assert COLOR_MAP["red"] == 1
        assert "blue" in COLOR_MAP
        assert COLOR_MAP["blue"] == 5
        assert "white" in COLOR_MAP
        assert COLOR_MAP["white"] == 7

    def test_normalize_coordinate_utility(self):
        """Test that CADInterface.normalize_coordinate works properly."""
        from core import CADInterface

        # Test 2D coordinate
        result_2d = CADInterface.normalize_coordinate((10.5, 20.5))
        assert result_2d == (10.5, 20.5, 0.0)

        # Test 3D coordinate
        result_3d = CADInterface.normalize_coordinate((10.5, 20.5, 30.5))
        assert result_3d == (10.5, 20.5, 30.5)


class TestLayersInfo:
    """Test suite for layer information extraction."""

    def test_get_layers_info_with_entity_data_parameter(self):
        """Test that get_layers_info accepts pre-extracted entity_data to avoid re-iteration."""
        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter = AutoCADAdapter("autocad")

        # Pre-extracted entity data
        entity_data = [
            {"Layer": "0", "Handle": "A1"},
            {"Layer": "0", "Handle": "A2"},
            {"Layer": "1", "Handle": "B1"},
            {"Layer": "Walls", "Handle": "C1"},
            {"Layer": "Walls", "Handle": "C2"},
            {"Layer": "Walls", "Handle": "C3"},
        ]

        with (
            patch.object(adapter, "_validate_connection"),
            patch.object(adapter, "_get_document") as mock_get_doc,
        ):
            # Mock document with layers
            mock_layer_0 = MagicMock()
            mock_layer_0.Name = "0"
            mock_layer_0.color = 7
            mock_layer_0.Linetype = "Continuous"
            mock_layer_0.Lineweight = -3
            mock_layer_0.Lock = False
            mock_layer_0.LayerOn = True

            mock_layer_1 = MagicMock()
            mock_layer_1.Name = "1"
            mock_layer_1.color = 1
            mock_layer_1.Linetype = "Continuous"
            mock_layer_1.Lineweight = -3
            mock_layer_1.Lock = False
            mock_layer_1.LayerOn = True

            mock_layer_walls = MagicMock()
            mock_layer_walls.Name = "Walls"
            mock_layer_walls.color = 5
            mock_layer_walls.Linetype = "Continuous"
            mock_layer_walls.Lineweight = -3
            mock_layer_walls.Lock = False
            mock_layer_walls.LayerOn = True

            mock_doc = MagicMock()
            mock_doc.Layers = [mock_layer_0, mock_layer_1, mock_layer_walls]
            mock_get_doc.return_value = mock_doc

            # Call get_layers_info with pre-extracted entity_data
            result = adapter.get_layers_info(entity_data=entity_data)

            assert isinstance(result, list)
            assert len(result) == 3

            # Check entity counts match pre-extracted data
            layer_0_info = next((layer for layer in result if layer["Name"] == "0"), None)
            assert layer_0_info is not None
            assert layer_0_info["ObjectCount"] == 2

            layer_1_info = next((layer for layer in result if layer["Name"] == "1"), None)
            assert layer_1_info is not None
            assert layer_1_info["ObjectCount"] == 1

            layer_walls_info = next((layer for layer in result if layer["Name"] == "Walls"), None)
            assert layer_walls_info is not None
            assert layer_walls_info["ObjectCount"] == 3

            # Verify ModelSpace was NOT accessed (optimization)
            assert not hasattr(mock_doc, "ModelSpace") or not mock_doc.ModelSpace.called

    def test_get_layers_info_without_entity_data_uses_selectionsets(self):
        """Test that get_layers_info uses SelectionSets if entity_data not provided."""
        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter = AutoCADAdapter("autocad")

        with (
            patch.object(adapter, "_validate_connection"),
            patch.object(adapter, "_get_document") as mock_get_doc,
        ):
            # Mock document with layers
            mock_layer_0 = MagicMock()
            mock_layer_0.Name = "0"
            mock_layer_0.color = 7
            mock_layer_0.Linetype = "Continuous"
            mock_layer_0.Lineweight = -3
            mock_layer_0.Lock = False
            mock_layer_0.LayerOn = True

            mock_layer_1 = MagicMock()
            mock_layer_1.Name = "1"
            mock_layer_1.color = 1
            mock_layer_1.Linetype = "Continuous"
            mock_layer_1.Lineweight = -3
            mock_layer_1.Lock = False
            mock_layer_1.LayerOn = True

            mock_doc = MagicMock()
            mock_doc.Layers = [mock_layer_0, mock_layer_1]

            # Mock SelectionSets
            mock_ss = MagicMock()
            # Set up count iteration: layer "0" has 1 object, layer "1" has 0 objects
            mock_ss.Count = 1

            # Let subsequent ss.Count calls return different values when needed.
            # Or just rely on ss.Count being 1 for all queries in this mock
            type(mock_ss).Count = PropertyMock(side_effect=[1, 0])

            mock_ss_coll = MagicMock()
            mock_ss_coll.Add.return_value = mock_ss
            mock_doc.SelectionSets = mock_ss_coll

            mock_get_doc.return_value = mock_doc

            # Mock win32com and pythoncom so windows imports do not fail or complain during tests

            mock_win32com = MagicMock()
            mock_pythoncom = MagicMock()
            mock_pythoncom.VT_ARRAY = 8192
            mock_pythoncom.VT_I2 = 2
            mock_pythoncom.VT_VARIANT = 12

            with patch.dict(
                "sys.modules", {"win32com.client": mock_win32com, "pythoncom": mock_pythoncom}
            ):
                # Call get_layers_info without entity_data
                result = adapter.get_layers_info()

            assert isinstance(result, list)
            assert len(result) == 2

            # Check entity counts from SelectionSets iteration
            layer_0_info = next((layer for layer in result if layer["Name"] == "0"), None)
            assert layer_0_info is not None
            assert layer_0_info["ObjectCount"] == 1

            layer_1_info = next((layer for layer in result if layer["Name"] == "1"), None)
            assert layer_1_info is not None
            assert layer_1_info["ObjectCount"] == 0


class TestBlockCreation:
    """Test suite for block creation functionality."""

    def test_create_block_from_entities_success(self):
        """Test successful block creation from entity handles."""
        adapter = AutoCADAdapter("autocad")

        # Mock COM objects
        mock_app = MagicMock()
        mock_doc = MagicMock()
        mock_blocks = MagicMock()
        mock_block_def = MagicMock()

        # Setup mocks
        mock_app.ActiveDocument = mock_doc
        mock_doc.Blocks = mock_blocks
        mock_blocks.Add.return_value = mock_block_def
        mock_blocks.Item.side_effect = Exception("Block doesn't exist")  # Block name is unique

        # Mock HandleToObject for entities
        mock_entity1 = MagicMock()
        mock_entity2 = MagicMock()
        mock_doc.HandleToObject.side_effect = [mock_entity1, mock_entity2]

        # Mock CopyObjects
        mock_doc.CopyObjects = MagicMock()

        with (
            patch.object(adapter, "_validate_connection"),
            patch.object(adapter, "_get_application", return_value=mock_app),
        ):
            result = adapter.create_block_from_entities(
                block_name="TestBlock",
                entity_handles=["A1B2", "C3D4"],
                insertion_point=(0.0, 0.0, 0.0),
                description="Test block",
            )

            # Verify result
            assert result["success"] is True
            assert result["block_name"] == "TestBlock"
            assert result["total_handles"] == 2
            assert result["entities_added"] == 2
            assert result["failed_handles"] == []
            assert result["insertion_point"] == (0.0, 0.0, 0.0)

            # Verify block was created
            mock_blocks.Add.assert_called_once()
            mock_doc.CopyObjects.assert_called_once()

    def test_create_block_from_entities_duplicate_name(self):
        """Test that creating block with existing name raises error."""
        adapter = AutoCADAdapter("autocad")

        # Mock COM objects
        mock_app = MagicMock()
        mock_doc = MagicMock()
        mock_blocks = MagicMock()
        mock_existing_block = MagicMock()

        # Setup mocks - block already exists
        mock_app.ActiveDocument = mock_doc
        mock_doc.Blocks = mock_blocks
        mock_blocks.Item.return_value = mock_existing_block  # Block exists (doesn't raise)

        from core.exceptions import CADOperationError

        with (
            patch.object(adapter, "_validate_connection"),
            patch.object(adapter, "_get_application", return_value=mock_app),
        ):
            with pytest.raises(CADOperationError) as exc_info:
                adapter.create_block_from_entities(
                    block_name="ExistingBlock",
                    entity_handles=["A1B2"],
                    insertion_point=(0.0, 0.0, 0.0),
                )

            assert "already exists" in str(exc_info.value).lower()

    def test_create_block_from_entities_invalid_handles(self):
        """Test block creation with invalid entity handles."""
        adapter = AutoCADAdapter("autocad")

        # Mock COM objects
        mock_app = MagicMock()
        mock_doc = MagicMock()
        mock_blocks = MagicMock()
        mock_block_def = MagicMock()

        # Setup mocks
        mock_app.ActiveDocument = mock_doc
        mock_doc.Blocks = mock_blocks
        mock_blocks.Add.return_value = mock_block_def
        mock_blocks.Item.side_effect = Exception("Block doesn't exist")

        # Mock HandleToObject - one entity exists, one doesn't
        mock_entity = MagicMock()
        mock_doc.HandleToObject.side_effect = [mock_entity, Exception("Invalid handle")]

        mock_doc.CopyObjects = MagicMock()

        with (
            patch.object(adapter, "_validate_connection"),
            patch.object(adapter, "_get_application", return_value=mock_app),
        ):
            result = adapter.create_block_from_entities(
                block_name="PartialBlock",
                entity_handles=["VALID", "INVALID"],
                insertion_point=(0.0, 0.0, 0.0),
            )

            # Should succeed with 1 entity, report 1 failed
            assert result["success"] is True
            assert result["entities_added"] == 1
            assert len(result["failed_handles"]) == 1
            assert "INVALID" in result["failed_handles"]

    def test_create_block_from_selection_success(self):
        """Test successful block creation from selected entities."""
        adapter = AutoCADAdapter("autocad")

        # Mock COM objects
        mock_app = MagicMock()
        mock_doc = MagicMock()
        mock_blocks = MagicMock()
        mock_block_def = MagicMock()
        mock_selection = MagicMock()

        # Setup mocks
        mock_app.ActiveDocument = mock_doc
        mock_doc.Blocks = mock_blocks
        mock_doc.PickfirstSelectionSet = mock_selection
        mock_blocks.Add.return_value = mock_block_def
        mock_blocks.Item.side_effect = Exception("Block doesn't exist")

        # Mock selection with 3 entities
        mock_entity1 = MagicMock()
        mock_entity2 = MagicMock()
        mock_entity3 = MagicMock()
        mock_selection.Count = 3
        mock_selection.Item.side_effect = [mock_entity1, mock_entity2, mock_entity3]

        mock_doc.CopyObjects = MagicMock()

        with (
            patch.object(adapter, "_validate_connection"),
            patch.object(adapter, "_get_application", return_value=mock_app),
        ):
            result = adapter.create_block_from_selection(
                block_name="SelectionBlock",
                insertion_point=(10.0, 20.0, 0.0),
                description="From selection",
            )

            # Verify result
            assert result["success"] is True
            assert result["block_name"] == "SelectionBlock"
            assert result["entities_added"] == 3
            assert result["insertion_point"] == (10.0, 20.0, 0.0)

            # Verify block was created
            mock_blocks.Add.assert_called_once()
            mock_doc.CopyObjects.assert_called_once()

    def test_create_block_from_selection_no_selection(self):
        """Test that error is raised when no entities are selected."""
        adapter = AutoCADAdapter("autocad")

        # Mock COM objects
        mock_app = MagicMock()
        mock_doc = MagicMock()
        mock_blocks = MagicMock()
        mock_selection = MagicMock()

        # Setup mocks - no entities selected
        mock_app.ActiveDocument = mock_doc
        mock_doc.Blocks = mock_blocks
        mock_doc.PickfirstSelectionSet = mock_selection
        mock_blocks.Item.side_effect = Exception("Block doesn't exist")
        mock_selection.Count = 0  # No selection

        from core.exceptions import CADOperationError

        with (
            patch.object(adapter, "_validate_connection"),
            patch.object(adapter, "_get_application", return_value=mock_app),
        ):
            with pytest.raises(CADOperationError) as exc_info:
                adapter.create_block_from_selection(
                    block_name="EmptyBlock", insertion_point=(0.0, 0.0, 0.0)
                )

            assert "No entities selected" in str(exc_info.value)

    def test_create_block_from_entities_invalid_name(self):
        """Test that invalid block name raises error."""
        adapter = AutoCADAdapter("autocad")

        from core.exceptions import InvalidParameterError

        # Mock COM objects
        mock_app = MagicMock()
        mock_doc = MagicMock()
        mock_app.ActiveDocument = mock_doc

        with (
            patch.object(adapter, "_validate_connection"),
            patch.object(adapter, "_get_application", return_value=mock_app),
        ):
            # Empty string
            with pytest.raises(InvalidParameterError):
                adapter.create_block_from_entities(block_name="", entity_handles=["A1B2"])

            # None
            with pytest.raises(InvalidParameterError):
                adapter.create_block_from_entities(block_name=None, entity_handles=["A1B2"])

    def test_objects_to_variant_array_helper(self):
        """Test _objects_to_variant_array helper method."""
        adapter = AutoCADAdapter("autocad")

        # Mock COM objects
        mock_obj1 = MagicMock()
        mock_obj2 = MagicMock()
        objects = [mock_obj1, mock_obj2]

        # Call helper
        result = adapter._objects_to_variant_array(objects)

        # Verify it returns a variant (we can't check exact type without pythoncom)
        assert result is not None


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
